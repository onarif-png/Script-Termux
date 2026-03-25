"""
ETF STOCK RANKER V4
================================
Ranks stocks by ETF coverage weighted by each ETF's Assets Under Management.

Every ETF manager vote counts equally — top 30 positions only.
This makes the ranking reflect real institutional conviction.

SCORING FORMULA:
  For each ETF holding a stock:
    points = (30 - rank + 1)  →  rank1=30pts, rank30=1pt

  points = (30 - rank + 1)
  (so median ETF = weight 1.0, VOO = weight ~200x, tiny ETF = weight ~0.01x)

HOW TO RUN:
  python3 etf_ranker_v4.py           — full scan (weekly)
  python3 etf_ranker_v4.py 100       — test first 100 ETFs
  python3 etf_ranker_v4.py SPY QQQ   — test specific ETFs

WEEKLY WORKFLOW:
  1. Download fresh screener CSV from stockanalysis.com/screener/etf
     Save to "CSV TICKER LISTS" subfolder
  2. Refresh cookies.json (Cookie-Editor export → pbpaste)
  3. Run: caffeinate -i python3 etf_ranker_v4.py

REQUIRES in same folder:
  - cookies.json
  - screener CSVs in "CSV TICKER LISTS" subfolder

OUTPUT:
  - etf_ranking_YYYY-MM-DD.xlsx   ← used by lookup.py for daily stock checks
  - etf_ranking_YYYY-MM-DD.csv
"""

import sys, requests, json, time, glob, re, warnings
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Paths ─────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).parent
CSV_DIR     = SCRIPT_DIR / "CSV TICKER LISTS"
DATE_STR    = datetime.now().strftime("%Y-%m-%d")
OUTPUT_XLSX = SCRIPT_DIR / f"etf_ranking_v4_{DATE_STR}.xlsx"
OUTPUT_CSV  = SCRIPT_DIR / f"etf_ranking_v4_{DATE_STR}.csv"
OUTPUT_MAP  = SCRIPT_DIR / f"etf_holdings_map_{DATE_STR}.csv"
DELAY       = 0.8
MAX_RETRIES = 3

# ── Cookie loader ─────────────────────────────────────────────────
def load_cookie_string():
    for p in [SCRIPT_DIR/"cookies.json", Path.home()/"Desktop"/"cookies.json"]:
        if p.exists():
            raw = json.loads(p.read_text())
            if isinstance(raw, list):
                return "; ".join(f"{c['name']}={c['value']}" for c in raw if "name" in c)
            return "; ".join(f"{k}={v}" for k,v in raw.items())
    return ""

# ── Load ETF list from screener CSV ────────────────────────────────
def load_etf_data():
    """
    Reads the screener CSV and returns:
      etf_list: [ticker, ...]  in original order
      aum_map:  {ticker: aum_in_dollars}
    """
    patterns = [
        str(CSV_DIR / "screener-etf-*.csv"), str(CSV_DIR / "etf-list.csv"),
        str(CSV_DIR / "etf_list.csv"),       str(CSV_DIR / "*.csv"),
        str(SCRIPT_DIR / "screener-etf-*.csv"), str(SCRIPT_DIR / "etf-list.csv"),
    ]
    csv_file = None
    for p in patterns:
        files = sorted(glob.glob(p))
        if files: csv_file = files[-1]; break

    if not csv_file:
        print("  ❌ No ETF CSV found in CSV TICKER LISTS folder!")
        print(f"     Download from stockanalysis.com/screener/etf")
        print(f"     Save to: {CSV_DIR}")
        return [], {}

    df = pd.read_csv(csv_file)
    if "Symbol" not in df.columns:
        print(f"  ❌ No Symbol column in {Path(csv_file).name}")
        return [], {}

    # Extract tickers
    tickers = df["Symbol"].dropna().str.strip().str.upper().tolist()
    tickers = [t for t in tickers if t and len(t) <= 6]

    # Extract AUM, holdings count, category for ETF metadata
    aum_map      = {}  # {sym: float}
    holdings_map = {}  # {sym: int}   — total holdings count
    category_map = {}  # {sym: str}   — ETF category/name
    aum_col = None
    for col in ["Assets", "AUM", "Total Assets", "Net Assets"]:
        if col in df.columns: aum_col = col; break

    for _, row in df.iterrows():
        sym = str(row.get("Symbol","")).strip().upper()
        if not sym: continue
        if aum_col:
            aum = row.get(aum_col)
            if aum and str(aum) not in ["nan","None",""]:
                try:
                    _a = str(aum).strip().replace("$","").replace(",","")
                    if   _a.endswith("T"): aum_map[sym] = float(_a[:-1]) * 1e12
                    elif _a.endswith("B"): aum_map[sym] = float(_a[:-1]) * 1e9
                    elif _a.endswith("M"): aum_map[sym] = float(_a[:-1]) * 1e6
                    elif _a.endswith("K"): aum_map[sym] = float(_a[:-1]) * 1e3
                    else:                  aum_map[sym] = float(_a)
                except: pass
        hold = row.get("Holdings")
        if hold and str(hold) not in ["nan","None",""]:
            try: holdings_map[sym] = int(float(hold))
            except: pass
        # Fund Name + Category for industry matching
        name = str(row.get("Fund Name","") or "").strip()
        cat  = str(row.get("Category","")  or "").strip()
        category_map[sym] = f"{name} {cat}".lower()

    if aum_map:
        print(f"  ✅ {len(tickers)} ETFs loaded from {Path(csv_file).name}")
        print(f"  ✅ AUM data: {len(aum_map)} ETFs  "
              f"(largest: {max(aum_map, key=aum_map.get)} "
              f"${max(aum_map.values())/1e9:.0f}B)")
    else:
        print(f"  ✅ {len(tickers)} ETFs loaded (no AUM column found)")

    return tickers, aum_map, holdings_map, category_map

# ── Load Dollar Volume ranks from stock screener CSV ─────────────
def load_stock_data():
    """
    Reads stock screener CSV (from stockanalysis.com/screener/stock)
    Requires columns: Symbol, Stock Price, Avg. Volume
    Optional column:  Price Target

    Returns:
      dv_ranks:      {symbol: rank}   ranked by Avg Volume × Price
      price_targets: {symbol: target} analyst price target
    """
    patterns = [
        str(CSV_DIR / "screener-stocks-*.csv"),
        str(CSV_DIR / "screener-stock-*.csv"),
        str(CSV_DIR / "stocks-list.csv"),
        str(CSV_DIR / "stock-list.csv"),
        str(CSV_DIR / "stocks.csv"),
    ]
    csv_file = None
    for p in patterns:
        files = sorted(glob.glob(p))
        if files: csv_file = files[-1]; break

    if not csv_file:
        print("  ⚠️  No stock screener CSV found — DV rank & price target will be skipped")
        print(f"     Download from stockanalysis.com/screener/stock")
        print(f"     Columns needed: Symbol, Stock Price, Avg. Volume, Price Target")
        print(f"     Save to: {CSV_DIR}")
        return {}, {}

    df = pd.read_csv(csv_file)
    print(f"  ✅ Stock screener: {len(df)} stocks from {Path(csv_file).name}")

    # ── Dollar Volume rank (Avg Volume × Price) ───────────────────
    dv_ranks = {}
    avg_vol_col   = next((c for c in df.columns if "avg" in c.lower() and "vol" in c.lower()), None)
    price_col     = next((c for c in df.columns if c.lower() in ["stock price","price","close"]), None)

    if avg_vol_col and price_col:
        df["_dv"] = pd.to_numeric(df[avg_vol_col], errors="coerce") *                     pd.to_numeric(df[price_col],   errors="coerce")
        df_dv = df.dropna(subset=["_dv"]).sort_values("_dv", ascending=False).reset_index(drop=True)
        for rank, (_, row) in enumerate(df_dv.iterrows(), 1):
            sym = str(row.get("Symbol","")).strip().upper()
            if sym: dv_ranks[sym] = rank
        top3 = list(dv_ranks.items())[:3]
        print(f"  ✅ DV rank (Avg Vol × Price): {len(dv_ranks)} stocks")
        print(f"     Top 3: {', '.join(f'{s}(#{r})' for s,r in top3)}")
    else:
        print(f"  ⚠️  Need 'Avg. Volume' + 'Stock Price' columns for DV rank")
        print(f"     Found columns: {list(df.columns)}")

    # ── Price targets ─────────────────────────────────────────────
    price_targets = {}
    pt_col = next((c for c in df.columns
                   if "target" in c.lower() or "pt" == c.lower()), None)

    if pt_col:
        for _, row in df.iterrows():
            sym = str(row.get("Symbol","")).strip().upper()
            pt  = row.get(pt_col)
            if sym and pt and str(pt) not in ["nan","None","","-"]:
                try:
                    price_targets[sym] = round(float(str(pt).replace("$","").replace(",","")), 2)
                except: pass
        print(f"  ✅ Price targets: {len(price_targets)} stocks have targets")
    else:
        print(f"  ⚠️  No price target column found — add 'Price Target' to screener download")

    # ── Industry / Sector map ─────────────────────────────────────
    stock_industry = {}  # {sym: "industry sector"}
    sec_col = next((c for c in df.columns if c.lower() == "sector"),   None)
    ind_col = next((c for c in df.columns if c.lower() == "industry"), None)
    if sec_col or ind_col:
        for _, row in df.iterrows():
            sym = str(row.get("Symbol","")).strip().upper()
            if sym:
                ind = str(row.get(ind_col,"") or "").strip() if ind_col else ""
                sec = str(row.get(sec_col,"") or "").strip() if sec_col else ""
                stock_industry[sym] = f"{ind} {sec}".lower()
        print(f"  ✅ Industry/sector: {len(stock_industry)} stocks mapped")

    return dv_ranks, price_targets, stock_industry


# ── Holdings parser ───────────────────────────────────────────────
def parse_holdings(json_data):
    try:
        for node in json_data.get("nodes", []):
            if not isinstance(node, dict): continue
            if node.get("type") != "data": continue
            data = node.get("data", [])
            if not isinstance(data, list) or not data: continue
            meta = data[0]
            if not isinstance(meta, dict) or "holdings" not in meta: continue
            holdings_list = data[meta["holdings"]]
            if not isinstance(holdings_list, list): continue
            rows = []
            for hi in holdings_list:
                if hi >= len(data): continue
                hmap = data[hi]
                if not isinstance(hmap, dict): continue
                try:
                    no_val = data[hmap["no"]] if "no" in hmap else len(rows)+1
                    n_val  = data[hmap["n"]]  if "n"  in hmap else ""
                    s_val  = data[hmap["s"]]  if "s"  in hmap else ""
                    symbol = str(s_val).replace("$","").strip().upper()
                    if not symbol or len(symbol) > 8: continue
                    rows.append({
                        "rank":   int(no_val) if no_val else len(rows)+1,
                        "symbol": symbol,
                        "name":   str(n_val),
                    })
                except: continue
            return rows if rows else None
    except: return None

def fetch_holdings(etf, session):
    url = f"https://stockanalysis.com/etf/{etf.lower()}/holdings/__data.json"
    for attempt in range(MAX_RETRIES):
        try:
            r = session.get(url, timeout=20)
            if r.status_code == 404: return None
            if r.status_code == 200: return parse_holdings(r.json())
            time.sleep(2 ** attempt)
        except:
            if attempt < MAX_RETRIES - 1: time.sleep(2 ** attempt)
    return None

# ── Excel builder ─────────────────────────────────────────────────
def build_excel(df, filepath):
    print("  Building Excel...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ETF Stock Ranking V3"
    cols = list(df.columns)

    # Section header colors
    section_colors = {
        "Rank": "1F2937", "Symbol": "1F2937", "Company_Name": "1F2937",
        "Score": "1A3A5C", "Raw_Score": "1A3A5C",
        "ETF_Count": "1A4731", "Avg_Rank": "1A4731",
        "Best_Rank": "2D4A1E", "BestRank_ETF": "2D4A1E",
        "ETF_Count_With_BestRank": "2D4A1E",
        
        "Top2_ETF1": "374151", "Top2_Rank1": "374151",
        "Top2_ETF2": "374151", "Top2_Rank2": "374151",
        "ETFs_With_BestRank": "1F2937",
        "Note": "1F2937",
    }

    # Header row
    for ci, col in enumerate(cols, 1):
        c = ws.cell(1, ci, col)
        c.fill = PatternFill("solid", fgColor=section_colors.get(col, "374151"))
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 35

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        rank = row.get("Rank", ri-1)
        if   rank <= 10:  bg = "FFF9C4"   # gold   — top 10
        elif rank <= 50:  bg = "E8F5E9"   # green  — top 50
        elif rank <= 200: bg = "EFF6FF"   # blue   — top 200
        elif rank <= 500: bg = "F9FAFB"   # light  — top 500
        elif ri % 2 == 0: bg = "F3F4F6"
        else:             bg = "FFFFFF"

        for ci, col in enumerate(cols, 1):
            val = row[col]
            if isinstance(val, float) and val == val: val = round(val, 2)
            elif isinstance(val, float): val = None
            c = ws.cell(ri, ci, val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(size=9, name="Calibri",
                         bold=(col in ["Rank", "Score", "Symbol"]))
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    widths = {
        "Rank": 6, "Symbol": 9, "Company_Name": 28,
        "Score": 13, "Raw_Score": 12,
        "ETF_Count": 10, "Avg_Rank": 10,
        "Best_Rank": 10, "BestRank_ETF": 13,
        "ETF_Count_With_BestRank": 14,
        
        "Top2_ETF1": 10, "Top2_Rank1": 10,
        "Top2_ETF2": 10, "Top2_Rank2": 10,
        "ETFs_With_BestRank": 50,
        "Note": 60,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 12)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    wb.save(filepath)
    print(f"  ✅ Saved → {Path(filepath).name}")

# ── MAIN ──────────────────────────────────────────────────────────
def main():
    print(f"\n{'='*65}")
    print(f"  ETF STOCK RANKER V4")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*65}\n")

    # Cookies
    cookie_str = load_cookie_string()
    if not cookie_str:
        print("  ❌ cookies.json not found!")
        print(f"     Save to: {SCRIPT_DIR}")
        return
    print(f"  ✅ Cookies loaded ({len(cookie_str.split(';'))} cookies)")

    # Load ETF list from screener CSV
    print()
    all_etf_list, aum_map, holdings_map, category_map = load_etf_data()
    if not all_etf_list: return

    # Load Dollar Volume ranks + price targets
    dv_ranks, price_targets, _stock_industry = load_stock_data()

    # Equal weight — every ETF manager vote counts the same

    # Apply test/manual mode
    args = sys.argv[1:]
    if len(args) == 1 and args[0].isdigit():
        etf_list = all_etf_list[:int(args[0])]
        print(f"\n  ✅ Test mode: first {len(etf_list)} ETFs")
    elif args:
        etf_list = [t.upper() for t in args]
        print(f"\n  ✅ Manual: {', '.join(etf_list)}")
    else:
        etf_list = all_etf_list
        print(f"\n  ✅ Full scan: {len(etf_list)} ETFs")

    # Session
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.0 Safari/605.1.15",
        "Accept":          "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer":         "https://stockanalysis.com/",
        "Cookie":          cookie_str,
    })

    # Quick test
    print("\n  Testing SPY...", end=" ", flush=True)
    test = fetch_holdings("SPY", session)
    if test: print(f"✅ {len(test)} holdings. Ready!\n")
    else: print("❌ Failed — refresh cookies.json"); return

    # ── Fetch all ETF holdings into memory ────────────────────────
    total   = len(etf_list)
    stocks  = {}   # {symbol: {name, etf_ranks, etf_sizes}}
    all_holdings = {}  # {etf: [(sym, rank), ...]} — ALL holdings for map
    success = 0
    start   = time.time()

    print(f"  Fetching {total} ETF holdings...\n")

    for i, etf in enumerate(etf_list, 1):
        elapsed = time.time() - start
        rate    = i / elapsed if elapsed > 0 else 1
        eta     = f"{int((total-i)/rate//60)}m{int((total-i)/rate%60)}s"
        print(f"  [{i:>4}/{total}]  {etf:<8}", end=" ", flush=True)

        holdings = fetch_holdings(etf, session)
        if holdings:
            actual_size = len(holdings)
            etf_size = 30  # fixed — rank1=30pts, rank30=1pt

            # Store ALL holdings for the map (no rank cap)
            all_holdings[etf] = [(h["symbol"], h["rank"]) for h in holdings if h.get("symbol")]

            for h in holdings:
                # Only count stocks in top 30 positions of each ETF
                if h["rank"] > 30:
                    continue
                sym = h["symbol"]
                if sym not in stocks:
                    stocks[sym] = {
                        "name":      h["name"],
                        "etf_ranks": {},
                        "etf_sizes": {},
                    }
                stocks[sym]["etf_ranks"][etf] = h["rank"]
                stocks[sym]["etf_sizes"][etf] = etf_size
                if h["name"] and not stocks[sym]["name"]:
                    stocks[sym]["name"] = h["name"]

            success += 1
            print(f"✅ {actual_size:>4} total ({min(actual_size,30)} counted)  stocks:{len(stocks):>5}  eta:{eta}")
        else:
            print(f"❌ skip  eta:{eta}")

        time.sleep(DELAY)

    elapsed_total = time.time() - start
    print(f"\n  ── Fetch complete ──────────────────────────────────────")
    print(f"  ETFs fetched : {success}/{total}")
    print(f"  Unique stocks: {len(stocks)}")
    print(f"  Time taken   : {int(elapsed_total//60)}m{int(elapsed_total%60)}s\n")

    # ── Build reverse map: ETF → [stocks sorted by rank] ─────────
    etf_top_stocks = {}  # {etf: ["SYM1","SYM2",...] sorted by rank}
    for sym, data in stocks.items():
        for etf, rank in data["etf_ranks"].items():
            if etf not in etf_top_stocks:
                etf_top_stocks[etf] = []
            etf_top_stocks[etf].append((rank, sym))
    for etf in etf_top_stocks:
        etf_top_stocks[etf] = [s for _, s in sorted(etf_top_stocks[etf])]

    # ── Rank stocks ───────────────────────────────────────────────
    print(f"  Ranking {len(stocks)} stocks by top-30 ETF coverage...")

    date_str = datetime.now().strftime("%-d%b%y")
    rows = []

    for sym, data in stocks.items():
        er = data["etf_ranks"]
        es = data["etf_sizes"]
        n  = len(er)

        # Score: equal weight per ETF — every manager's conviction counts the same
        score = sum((es.get(e, 30) - r + 1) for e, r in er.items())
        raw_score = score

        avg_rank  = round(sum(er.values()) / n, 2) if n else 0
        best_rank = min(er.values()) if er else 999
        best_etfs = sorted([e for e, r in er.items() if r == best_rank])

        # ── 3 ETF signals ─────────────────────────────────────────
        # 1. low.h — among best-rank ETFs: lowest holdings count (most focused)
        lowhld_etf = min(
            best_etfs,
            key=lambda e: holdings_map.get(e, 9999)
        ) if best_etfs else ""
        lowhld_rank = best_rank

        # 2. hi.a — among best-rank ETFs: highest AUM (most institutional)
        hiaum_etf = max(
            best_etfs,
            key=lambda e: aum_map.get(e, 0)
        ) if best_etfs else ""
        hiaum_rank = best_rank

        # 3. ind/sec — best ETF tag using holdings map
        stock_industry = _stock_industry.get(sym, "")
        ind_etf = ""; ind_rank = ""; ind_tag_prefix = "ind"
        if stock_industry and all_holdings:
            parts = stock_industry.split()
            stock_sector = parts[-1] if parts else ""

            # Industry match: best concentration %
            same_ind = {s for s, info in _stock_industry.items()
                        if info.lower() == stock_industry.lower()}
            min_pct = max(0.03, min(0.10, 3 / max(len(same_ind), 1)))
            top3_matches = []
            for e, pairs in all_holdings.items():
                total_h = len(pairs)
                if total_h < 5: continue
                mc = sum(1 for s, _ in pairs if s in same_ind)
                pct = mc / total_h
                if pct >= min_pct:
                    top3_matches.append((e, pct))
            top3_matches.sort(key=lambda x: -x[1])
            top3_matches = top3_matches[:3]
            if top3_matches:
                best_e = max(top3_matches, key=lambda x: aum_map.get(x[0], 0))
                ind_etf  = best_e[0]
                ind_rank = er.get(ind_etf, "-")

            # Sector fallback: highest-ranked sector ETF that holds this stock
            if not ind_etf and stock_sector:
                ind_tag_prefix = "sec"
                best_rank_val = 9999
                for e, pairs in all_holdings.items():
                    cat = category_map.get(e, "")
                    if stock_sector not in cat: continue
                    syms_in_etf = {s for s, _ in pairs}
                    if sym not in syms_in_etf: continue
                    r = er.get(e)
                    if r is None: continue
                    try:
                        r_int = int(r)
                    except: continue
                    if r_int < best_rank_val:
                        best_rank_val = r_int
                        ind_etf  = e
                        ind_rank = r

        dv_rank = dv_ranks.get(sym)
        dv_tag  = f" [DV-{dv_rank}]" if dv_rank else ""
        pt      = price_targets.get(sym)
        pt_tag  = f" (to {pt})" if pt else ""

        # Top 5 stocks from ind ETF (excluding current stock itself)
        ind_top5 = ""
        if ind_etf:
            ind_stocks = [s for s in etf_top_stocks.get(ind_etf, []) if s != sym][:5]
            if ind_stocks:
                ind_top5 = " " + ", ".join(ind_stocks)

        etf3 = (f"[low.h-{lowhld_etf}(R{lowhld_rank})"
                f" {len(best_etfs)}]"
                f"[hi.a-{hiaum_etf}(R{hiaum_rank})]"
                f"[{ind_tag_prefix}-{ind_etf}(R{ind_rank})]" if ind_etf else
                f"[low.h-{lowhld_etf}(R{lowhld_rank})"
                f" {len(best_etfs)}]"
                f"[hi.a-{hiaum_etf}(R{hiaum_rank})]")

        note = (f"{sym}-R{{RANK}}{dv_tag} – {n}ETF (Avg {avg_rank}) "
                f"{etf3}"
                f"{pt_tag}{ind_top5} {date_str}")

        rows.append({
            "Symbol":                  sym,
            "Company_Name":            data["name"],
            "Score":                   round(score, 1),
            "Raw_Score":               raw_score,
            "ETF_Count":               n,
            "Avg_Rank":                avg_rank,
            "Best_Rank":               best_rank,
            "BestRank_ETF":            lowhld_etf,
            "ETF_Count_With_BestRank": len(best_etfs),
            "Top2_ETF1":               hiaum_etf,
            "Top2_Rank1":              hiaum_rank,
            "Top2_ETF2":               ind_etf,
            "Top2_Rank2":              ind_rank,
            "ETFs_With_BestRank":      ", ".join(best_etfs),
            "Note":                    note,
        })

    # Sort by Score
    df = pd.DataFrame(rows)
    df = df.sort_values("Score", ascending=False).reset_index(drop=True)
    df.insert(0, "Rank", range(1, len(df)+1))
    df["Note"] = df.apply(
        lambda r: r["Note"].replace("{RANK}", str(int(r["Rank"]))), axis=1)

    # Print top 15
    print(f"\n  Top 15 by ETF coverage (top-30 equal weight):")
    print(f"  {'Rank':<6}{'Symbol':<10}{'Score':>12}{'RawScore':>10}"
          f"{'ETFs':>6}{'AvgRank':>9}  Best ETF")
    print(f"  {'─'*65}")
    for _, r in df.head(15).iterrows():
        print(f"  {int(r['Rank']):<6}{r['Symbol']:<10}"
              f"{r['Score']:>12,.0f}{int(r['Raw_Score']):>10}"
              f"{int(r['ETF_Count']):>6}{r['Avg_Rank']:>9.1f}"
              f"  {r['BestRank_ETF']}(#{r['Best_Rank']})")

    # Save outputs
    df.to_csv(OUTPUT_CSV, index=False)
    print(f"\n  ✅ CSV  → {OUTPUT_CSV.name}")
    build_excel(df, str(OUTPUT_XLSX))

    # ── Save full ETF→Stock holdings map (ALL holdings, not just top 30) ─
    map_rows = []
    for etf, pairs in all_holdings.items():
        for sym, rank in pairs:
            map_rows.append({"ETF": etf, "Symbol": sym, "Rank": rank})
    map_df = pd.DataFrame(map_rows).sort_values(["ETF", "Rank"]).reset_index(drop=True)
    map_df.to_csv(OUTPUT_MAP, index=False)
    print(f"  ✅ MAP  → {OUTPUT_MAP.name}  ({len(map_df):,} ETF-stock pairs)")

    print(f"\n{'='*65}")
    print(f"  DONE")
    print(f"  {len(df)} stocks ranked | {success} ETFs processed")
    print(f"  lookup.py will auto-read: {OUTPUT_XLSX.name}")
    print(f"{'='*65}\n")

if __name__ == "__main__":
    main()
